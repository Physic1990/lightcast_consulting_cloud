import openpyxl
import tkinter as tk
from tkinter import ttk
import time
from typing import List, Optional

class AggregationUI:
    def __init__(self, title: str = "Progress"):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.geometry("300x150")
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            self.root, 
            variable=self.progress_var,
            maximum=100
        )
        self.progress.pack(pady=10, padx=10, fill=tk.X)
        
        # Progress text
        self.text_var = tk.StringVar()
        self.text_label = ttk.Label(self.root, textvariable=self.text_var)
        self.text_label.pack(pady=5)
        
        # Update text
        self.update_var = tk.StringVar()
        self.update_label = ttk.Label(self.root, textvariable=self.update_var)
        self.update_label.pack(pady=5)
        
    def update_progress(self, progress: float, message: str):
        """Update progress bar and message"""
        self.progress_var.set(progress)
        self.text_var.set(message)
        self.root.update()
        
    def update_status(self, message: str):
        """Update status message"""
        self.update_var.set(message)
        self.root.update()
        
    def close(self):
        """Close the window"""
        self.root.destroy()

class ExcelAggregation:
    def __init__(self, workbook_path: str):
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)
        self.ins_sheet = self.wb["Ins"]
        self.ui_sheet = self.wb["UI"]
        self.agg_vector_sheet = self.wb["Agg Vector"]
        
    def get_states(self) -> List[str]:
        """Equivalent to GetStates() in VBA"""
        states = []
        # Loop through columns D to IU (4 to 256)
        for col in range(4, 257):
            state = self.ui_sheet.cell(row=9, column=col).value
            if state:
                states.append(state)
        return states
    
    def state_agg(self, state: str, main_process_func):
        """Equivalent to StateAgg() in VBA"""
        # Create progress UI
        progress_ui = AggregationUI("State Aggregation Progress")
        
        # Count selections
        num_selections = sum(
            1 for col in range(4, 257)
            if self.ui_sheet.cell(row=9, column=col).value == state
        )
        
        # Clear aggregation area
        for row in range(8, 2219):
            for col in range(20, 120):
                self.ins_sheet.cell(row=row, column=col).value = None
        
        column = 20
        
        # Process each matching institution
        for i in range(4, 257):
            if self.ui_sheet.cell(row=9, column=i).value == state:
                # Update progress
                progress = (column - 20) * 100 / num_selections
                message = f"Completed {column - 20}/{num_selections} institutions"
                progress_ui.update_progress(progress, message)
                
                # Set institution index
                self.ins_sheet["D3"] = i - 3
                
                # Run main process
                main_process_func()
                
                # Copy results
                for row in range(8, 5009):
                    self.ins_sheet.cell(row=row, column=column).value = \
                        self.agg_vector_sheet.cell(row=row-6, column=3).value
                
                column += 1
        
        # Final progress update
        progress_ui.update_progress(100, f"Completed {num_selections}/{num_selections} institutions")
        progress_ui.update_status("Done!")
        time.sleep(1)  # Give user time to see completion
        progress_ui.close()
        
    def custom_agg(self, selected_indices: List[int], main_process_func):
        """Equivalent to CustomAgg() in VBA"""
        # Create progress UI
        progress_ui = AggregationUI("Custom Aggregation Progress")
        
        num_selections = len(selected_indices)
        column = 20
        
        # Clear aggregation area
        for row in range(8, 2219):
            for col in range(20, 120):
                self.ins_sheet.cell(row=row, column=col).value = None
        
        # Process each selected institution
        for idx, i in enumerate(selected_indices):
            # Update progress
            progress = idx * 100 / num_selections
            message = f"Completed {idx}/{num_selections} institutions"
            progress_ui.update_progress(progress, message)
            
            # Set institution index
            self.ins_sheet["D3"] = i
            
            # Run main process
            main_process_func()
            
            # Copy results
            for row in range(8, 5009):
                self.ins_sheet.cell(row=row, column=column).value = \
                    self.agg_vector_sheet.cell(row=row-6, column=3).value
            
            column += 1
        
        # Final progress update
        progress_ui.update_progress(100, f"Completed {num_selections}/{num_selections} institutions")
        progress_ui.update_status("Done!")
        time.sleep(1)  # Give user time to see completion
        progress_ui.close()

# Example usage:
def main():
    # Initialize Excel automation
    excel = ExcelAggregation("path_to_your_workbook.xlsx")
    
    # Get list of states
    states = excel.get_states()
    print(f"Available states: {states}")
    
    # Example of running state aggregation
    def main_process():
        # Implementation of Main.MainProcess would go here
        pass
    
    # Run aggregation for a specific state
    excel.state_agg("California", main_process)
    
    # Run custom aggregation for selected indices
    selected_indices = [1, 3, 5]  # Example indices
    excel.custom_agg(selected_indices, main_process)

if __name__ == "__main__":
    main()
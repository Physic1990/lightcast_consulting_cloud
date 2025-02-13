import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def change_colors(workbook_path, sheet_name='Sheet1'):
    """
    Change font colors of specific cells in Excel worksheet
    """
    # Load workbook and select active sheet
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]
    
    # Get RGB values and ensure they're within 0-255 range
    l_red = abs(ws['K2'].value) % 256
    l_green = abs(ws['K3'].value) % 256
    l_blue = abs(ws['K4'].value) % 256
    
    # Second set of RGB values
    red2 = abs(ws['L2'].value) % 256
    green2 = abs(ws['L3'].value) % 256
    blue2 = abs(ws['L4'].value) % 256
    
    # First color set cell ranges
    color1_cells = [
        'B2',
        'E14', 'E15', 'E16',
        'F19', 'F20', 'F21', 'F22', 'F23', 'F24',
        'F41', 'F42', 'F43', 'F44', 'F45', 'F46', 'F47', 'F48', 'F49', 'F50',
        'F54', 'F55',
        'G25', 'G26', 'G27', 'G28'
    ]
    
    # Second color set cell ranges
    color2_cells = [
        'E17', 'E18',
        'E52', 'E53',
        'H30', 'H31',
        'H34', 'H35',
        'H38'
    ]
    
    # Apply first color set
    color1 = f"{l_red:02x}{l_green:02x}{l_blue:02x}"
    for cell_ref in color1_cells:
        cell = ws[cell_ref]
        cell.font = Font(color=color1)
    
    # Apply second color set
    color2 = f"{red2:02x}{green2:02x}{blue2:02x}"
    for cell_ref in color2_cells:
        cell = ws[cell_ref]
        cell.font = Font(color=color2)
    
    # Save the workbook
    wb.save(workbook_path)

# Alternative implementation using win32com for better compatibility
def change_colors_win32(workbook_path):
    """
    Change font colors using win32com (more compatible with Excel)
    """
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    
    try:
        # Open workbook
        wb = excel.Workbooks.Open(workbook_path)
        ws = wb.ActiveSheet
        
        # Get RGB values
        l_red = abs(ws.Range("K2").Value) % 256
        l_green = abs(ws.Range("K3").Value) % 256
        l_blue = abs(ws.Range("K4").Value) % 256
        
        red2 = abs(ws.Range("L2").Value) % 256
        green2 = abs(ws.Range("L3").Value) % 256
        blue2 = abs(ws.Range("L4").Value) % 256
        
        # First color set
        color1_cells = [
            "B2",
            "E14", "E15", "E16",
            "F19", "F20", "F21", "F22", "F23", "F24",
            "F41", "F42", "F43", "F44", "F45", "F46", "F47", "F48", "F49", "F50",
            "F54", "F55",
            "G25", "G26", "G27", "G28"
        ]
        
        # Second color set
        color2_cells = [
            "E17", "E18",
            "E52", "E53",
            "H30", "H31",
            "H34", "H35",
            "H38"
        ]
        
        # Apply colors
        for cell_ref in color1_cells:
            ws.Range(cell_ref).Font.Color = rgb_to_int(l_red, l_green, l_blue)
            
        for cell_ref in color2_cells:
            ws.Range(cell_ref).Font.Color = rgb_to_int(red2, green2, blue2)
        
        # Save and close
        wb.Save()
        wb.Close()
    finally:
        excel.Quit()

def rgb_to_int(red, green, blue):
    """Convert RGB values to integer color value for win32com"""
    return red + (green * 256) + (blue * 256 * 256)